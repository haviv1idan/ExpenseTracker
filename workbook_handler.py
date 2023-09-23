import json
import requests

from datetime import datetime
from openpyxl import load_workbook


filename = "transaction-details_export_1695383757436.xlsx"
api_url = "http://127.0.0.1:8000/api/"

# Load the workbook
workbook = load_workbook(filename)
 
# Iterate through all sheets
url = api_url + "transactions/create/"
 
    # Access the data in the sheet
for i, row in enumerate(workbook.active.iter_rows(values_only=True)):
    print(row)
    if i < 4 or row[14] is None:
        continue

    data = {
        "date": row[0],
        "business_name": row[1],
        "category": row[2],
        "credit_4_last_digits": row[3],
        "expense_type": row[4],
        "transaction_amount": row[5],
        "debit_currency": row[6],
        "original_transaction_amount": row[7],
        "original_transaction_currency": row[8],
        "billing_date": row[9], 
        "payment_type": row[14]
    }
    try:
        response = requests.post(url, data=data)
    except ValueError as e:
        print(e)
    if response.status_code != 201:
        print(response.status_code)
        print(response.text)
        print("#########################################")
